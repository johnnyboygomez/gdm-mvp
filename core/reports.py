# core/reports.py

from datetime import date, timedelta, datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .models import Participant


def generate_research_excel(participant_id=None):
    """
    Generate Excel file with research data - one row per week per participant.
    
    Args:
        participant_id: If provided, export only this participant. Otherwise export all.
    
    Returns:
        HttpResponse with Excel file
    """
    
    # Get participants (exclude staff and superuser)
    if participant_id:
        participants = Participant.objects.select_related('user').filter(
            id=participant_id,
            user__is_staff=False,
            user__is_superuser=False
        )
    else:
        participants = Participant.objects.select_related('user').filter(
            user__is_staff=False,
            user__is_superuser=False
        ).order_by('id')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Research Data"
    
    # Define headers
    headers = [
        'subject_ID',
        'tx_arm',
        'start_date',
        'week_number',
        'week_start_date',
        'avg_steps_per_day',
        'days_with_data',
        'total_steps',
        'previous_week_target',
        'reached_goal',
        'increment',
        'new_target',
    ]
    
    # Write headers with formatting
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    current_row = 2
    
    # Process each participant
    for participant in participants:
        daily_steps = participant.daily_steps or []
        targets = participant.targets or {}
        
        if not daily_steps:
            continue  # Skip participants with no data
        
        # Sort daily steps by date
        sorted_steps = sorted(daily_steps, key=lambda x: x.get('date', ''))
        
        # Group steps by week
        weeks_data = {}
        
        for step_entry in sorted_steps:
            step_date = step_entry.get('date')
            step_value = step_entry.get('value', 0)
            
            if not step_date:
                continue
            
            step_date_obj = date.fromisoformat(step_date)
            days_since_start = (step_date_obj - participant.start_date).days
            
            # Determine which week this day belongs to
            week_number = (days_since_start // 7) + 1
            
            if week_number not in weeks_data:
                weeks_data[week_number] = {
                    'steps': [],
                    'week_start': participant.start_date + timedelta(days=(week_number - 1) * 7),
                    'week_end': participant.start_date + timedelta(days=(week_number - 1) * 7 + 6)
                }
            
            weeks_data[week_number]['steps'].append(step_value)
        
        # Write one row per week
        for week_num in sorted(weeks_data.keys()):
            week_info = weeks_data[week_num]
            week_start = week_info['week_start']
            week_end = week_info['week_end']
            week_steps = week_info['steps']
            
            # Calculate week statistics
            days_with_data = len(week_steps)
            total_steps = sum(week_steps)
            avg_steps = round(total_steps / days_with_data) if days_with_data > 0 else 0
            
            # Get target data for this week
            week_key = week_start.strftime("%Y-%m-%d")
            target_data = targets.get(week_key, {})
            
            # Get previous week's target
            if week_num > 1:
                prev_week_start = week_start - timedelta(days=7)
                prev_week_key = prev_week_start.strftime("%Y-%m-%d")
                prev_target_data = targets.get(prev_week_key, {})
                prev_target = prev_target_data.get('new_target', None)
            else:
                prev_target = None
            
			# Determine if goal was reached
            if week_num == 1:
                reached_goal = 'NA'
            elif prev_target and target_data:
                average_from_target = target_data.get('average_steps', avg_steps)
                # Ensure both values are integers for comparison
                try:
                    avg_int = int(average_from_target)
                    target_int = int(prev_target)
                    reached_goal = 'Yes' if avg_int >= target_int else 'No'
                except (ValueError, TypeError):
                    reached_goal = 'NA'
            elif not target_data:
                reached_goal = '4'  # Not enough valid dates
            else:
                reached_goal = 'NA'            
            # Get increment
            if target_data:
                increment = target_data.get('increase', '')
            else:
                increment = ''
            
            # Get new target
            if target_data and target_data.get('new_target'):
                new_target = target_data['new_target']
            elif week_num > 1 and prev_target:
                # Carry forward previous target if no new one calculated
                new_target = prev_target
            else:
                new_target = ''
            
            # Write row
            ws.cell(row=current_row, column=1, value=participant.id)
            ws.cell(row=current_row, column=2, value=participant.treatment_arm)
            
            # start_date only on first week for this participant
            if week_num == 1:
                ws.cell(row=current_row, column=3, value=participant.start_date)
            
            ws.cell(row=current_row, column=4, value=week_num)
            ws.cell(row=current_row, column=5, value=week_start)  # Show every week
            ws.cell(row=current_row, column=6, value=avg_steps)
            ws.cell(row=current_row, column=7, value=days_with_data)
            ws.cell(row=current_row, column=8, value=total_steps)
            ws.cell(row=current_row, column=9, value=prev_target if prev_target else 'NA')
            ws.cell(row=current_row, column=10, value=reached_goal)
            ws.cell(row=current_row, column=11, value=increment)
            ws.cell(row=current_row, column=12, value=new_target)
            
            current_row += 1
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Create Dictionary sheet
    ws_dict = wb.create_sheet("Dictionary")
    dict_headers = ['variable', 'Plain English', 'Choices', 'Note']
    
    for col_num, header in enumerate(dict_headers, 1):
        cell = ws_dict.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Add dictionary entries
    dictionary_data = [
        ('subject_ID', 'Participant ID number', 'unique identifier', None),
        ('tx_arm', 'treatment arm', '0=Control | 1=Intervention', 'Does not change week to week'),
        ('start_date', 'First day the Fitbit was used', 'YYYY-MM-DD format', 'Shown only on first row per participant'),
        ('week_number', 'Study week number', '1, 2, 3, ...', 'Week 1 is baseline'),
        ('week_start_date', 'First day of this week', 'YYYY-MM-DD format', 'Shown for every week'),
        ('avg_steps_per_day', 'Average daily steps for this week', 'whole number', 'Rounded to nearest integer'),
        ('days_with_data', 'Number of days with step data', '0-7', 'Days with synced Fitbit data'),
        ('total_steps', 'Total steps for the week', 'whole number', 'Sum of all daily steps'),
        ('previous_week_target', 'Target that was set for this week', 'steps/day', 'Set at end of previous week'),
        ('reached_goal', 'Did participant reach the target?', 'Yes | No | NA | 4', 'NA for week 1; 4 = insufficient data'),
        ('increment', 'How target was adjusted', '+250, +500, +1000, maintain, etc.', 'Based on algorithm'),
        ('new_target', 'Target set for next week', 'steps/day', 'Calculated at end of this week'),
    ]
    
    for row_num, (var, plain, choices, note) in enumerate(dictionary_data, 2):
        ws_dict.cell(row=row_num, column=1, value=var)
        ws_dict.cell(row=row_num, column=2, value=plain)
        ws_dict.cell(row=row_num, column=3, value=choices)
        ws_dict.cell(row=row_num, column=4, value=note)
    
    # Auto-size dictionary columns
    for col in range(1, len(dict_headers) + 1):
        ws_dict.column_dimensions[get_column_letter(col)].width = 30
    
    # Prepare response
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    if participant_id:
        filename = f'partnersteps_participant_{participant_id}_{timestamp}.xlsx'
    else:
        filename = f'partnersteps_research_data_{timestamp}.xlsx'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response